# FileName : Excel.py
# Author   : Adil
# DateTime : 2017/12/10 13:08
# SoftWare : PyCharm

import xlrd
import os
from datetime import datetime
from openpyxl.reader.excel import load_workbook
from openpyxl import Workbook
from CiTest import readConfig as RC

Rc = RC.ReadConfig()

class Excel(object):
    '''定义一个excel类'''
    # 定义类变量
    i = 0
    def __init__(self):
        '初始化基本信息'

        self.path = Rc.path
        self.readExcelPath = os.path.join(self.path,'caseData')
        self.reportPath = os.path.join(self.readExcelPath, 'caseReport')
        # 创建caseReport 目录
        if not os.path.exists(self.reportPath):
            os.mkdir(self.reportPath)
        self.reportDatePath = os.path.join(self.reportPath, str(datetime.now().strftime('%Y%m%d')))
        # 创建reportDatePath 目录
        if not os.path.exists(self.reportDatePath):
            os.mkdir(self.reportDatePath)
        self.writeExcelFile = 'ApiReport-' + str(datetime.now().strftime('%H%M%S')) + '.xlsx'
        self.writeExcelName = os.path.join(self.reportDatePath, self.writeExcelFile)
        self.createExcel(self.writeExcelName)



    def createExcel(self,excelName):
        '''创建excel'''
        wb = Workbook()
        wb.save(excelName)

    def readExcel(self,excelName,SheetName):
        '读取excel'
        self.excelName = os.path.join(self.readExcelPath,excelName)
        self.Rb = xlrd.open_workbook(self.excelName)
        self.Rs = self.Rb.sheet_by_name(SheetName)
        # 获取行数
        rows = self.Rs.nrows
        # 定义一个dict存放单条用例
        # self.titleDict = dict.fromkeys(self.Rs.row_values(0))
        # 取第一行的表头存为list。
        self.titleList = self.Rs.row_values(0)
        # 定义一个list 存放 所有用例
        self.caseList = []
        self.tempList = []
        for r in range(1,rows):
            rowValues = self.Rs.row_values(r)
            # print(r)
            # print(self.Rs.row_values(r))
            # self.caseInfo = dict.fromkeys(self.Rs.row_values(0),self.Rs.row_values(r))
            # print(self.caseInfo)
            # 将列表组合成 字典 这是 将列表转换为字典的一个方法。
            self.caseDict = dict(zip(self.titleList,rowValues))
            # 下面是将字典转换为列表，
            # print(list(self.caseDict))
            # print(self.caseDict.values())
            # print(self.caseDict)
            # 将字典再拼接为列表。
            self.caseList.append(self.caseDict)
            self.tempList.append(rowValues)
        # print(self.caseList)
        # 返回caseList
        # self.writeExcel(SheetName, self.titleList, self.tempList)
        return self.caseList


    def writeExcel(self,SheetName,titleList,dataList):
        '''写入excel'''
        wb = load_workbook(self.writeExcelName)
        # 引用类变量 作为index
        sheetIndex = Excel.i
        # wb = load_workbook(self.writeExcelName)
        # 以SheetName 新建一个sheet页。 sheetIndex 作为序号
        ws = wb.create_sheet(SheetName,index=sheetIndex)
        ws.append(titleList)
        for dataDict in dataList:
            # titleList = list(dataDict)
            resultList = list(dataDict.values())
            ws.append(resultList)
        # 自增
        Excel.i += 1
        wb.save(filename=self.writeExcelName)

if __name__ == '__main__':

    excel = Excel()
    excel.readExcel('ApiInfo.xlsx','Login')